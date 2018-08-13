# Excel file manager for workscope exhibit generator tool
#
# NOTES: 
#   1. This module was written to run under Python 2.7.x
#   2. This module relies upon the OpenPyXl library being installed
#       OpenPyXl is used to read and navigate the input .xlsx workbook
#   3. To install OpenPyXl under Python 2.7.x:
#       <Python_installation_folder>/python.exe -m pip install openpyxl
#
# The code in this module previously resided in 'workscope_exhibit_tool.py'
#
# Author: Benjamin Krepp
# Date: 23-27 July, 30-31 July, 6-10 August 2018
#
# Requirements on the input .xlsx spreadsheet
# ===========================================
#
# This module relies upon the workscope exhibit template .xlsx file having been created
# with 'defined_names' for various locations of interest in it.
# Although 'defined names' are often used to specify ranges of cells, here they are used
# to identify specific cells, whose row and/or column index is obtained for use in the
# generation of the HTML for a workscope exhibit.
#
# 1. The worksheet containing the workscope exhibits MUST be named 'workscope_exhibits'.
#    Other worksheets may be present; their contents are ignored by this script.
#
# 2. These 'defined names' MUST be present in the workbook and be defined as folllows:
#
#   project_name_cell - cell containing the project's name
#   direct_salary_cell - cell containing the total direct salary and overhead
#   odc_cell - cell containing the total of other direct costs
#   total_cost_cell - cell containing the total cost of the project
#   task_list_top - any cell in line immediately preceeding list of tasks;
#                   only the row index of this cell is used
#   task_list_bottom - any cell in line immediately following list of tasks;
#                      only the row index of this cell is used
#   funding list_top - any cell in line immediately preceeding list of funding
#                      sources; only the row index of this cell is used
#   funding list_bottom - any cell in line immediately following list of funding
#                         sources; only the row index of this cell is used
#   task_number_column - any cell in column containing the task numbers in
#                        the cost table; only the column index of this cell is used
#   task_name_column - any cell in column containint the task name in the 
#                      cost table; only the column index of this cell is used
#   m1_column -    cost table header cell containing the text 'M-1'
#   p5_column -    cost table header cell containing the text 'P-5'
#   p4_column -    cost table header cell containing the text 'P-4' 
#   p3_column -    cost table header cell containing the text 'P-3'
#   p2_column -    cost table header cell containing the text 'P-2'
#   p1_column -    cost table header cell containing the text 'P-1'
#   sp3_column -   cost table header cell containing the text 'SP-3'
#   sp1_column -   cost table header cell containing the text 'SP-1'
#   temp_column -  cost table header cell containing the text 'Temp'
#   total_column - cost table header cell containing the text 'Total'
#   direct_salary_column - cost table header cell containg the string
#                          'Salary' (2nd line of 'Direct Salary')
#   overhead_column - cost table header cell containing the overhead
#                     rate (as a string); used when we want to access the column
#   overhead_cell -   ditto; used when we want to access the cell itself
#   total_cost_column - cost table header cell containing the text
#                       'Cost' (2nd line of 'Total Cost'
#   total_line - any cell in row containing person-hour totals in 
#                the cost table; only the row index of this cell is used
#   odc_travel_line - any cell on line containing Other Direct Costs: travel;
#                     only the row index of this cell is used
#   odc_office_equipment_line - any cell on line containing Other Direct Costs:
#                               general office equipment; only the row index of 
#                               this cell is used
#   odc_dp_equipment_line - any cell on line containing Other Direct Costs:
#                           data processing equipment; only the row index of 
#                           this cell is used
#   odc_consultants_line - any cell on line containing Other Direct Costs:
#                          consultants; only the row index of this cell is used
#   odc_printing_line - any cell on line containing Other Direct Costs:
#                       printing; only the row index of this cell is used
#   odc_other_line - any cell on line containing Ohter Direct Costs: other;
#                    only the row index of this cell is used
#   sched_major_units_cell - cell containing the 'major units' (which may be
#                       Quarters, Months, or Weeks) used to express the schedule
#
# Internals of this Module: Top-level Functions
# =============================================
#
# initExcelFile - Reads a completed .xlsx workscope exhibit template; extracts the row-
#                 and colum-indices (and a couple of other things) of interest/use, 
#                 which are stored in a dictionary object. This object is subsequently
#                 used throughout the 'workscope_exhibit_tool.py' module; it is the most 
#                 important data structure in the program as a whole.
#
# dump_xlsInfo - Dumps contents of data structure generated by initExcelFile in 
#                human-readable format.
#
# get_sched_col_info - FOSSIL TO BE REMOVED
#
# Internals of this Module: Utility Functions
# ===========================================
#
# get_column_index - return the column index for a defined name assigned to a single cell
#
# get_row_index - return the row index for a defined name assigned to a single cell
#
# get_cell_contents - returns the contents of a cell, given a worksheet name, row index,
#                     and column index
#
#
# A Guide for the Perplexed (with apologies to Maimonides),
#                           or
# An 'Ultra-quick' Quick-start Guide to Using the OpenPyXl Library
# ================================================================
#
# Open an .xlsx workbook:
#   wb = openpyxl.load_workbook(full_path_to_workbook_file, data_only=True)
#
# Get list of worksheets in workbook:
#   ws_list = wb.sheetnames
#
# Get a named worksheet:
#   ws = wb['workscope_template']
#
# Get list of defined names in workbook:
#   list_of_dns = wb.defined_names
#
# Get value of a defined name, e.g., 'foobar'
#   dn_val = wb.defined_names['foobar'].value
#
# Get the worksheet and cell indices for a defined name,
# and get the value of the cell it refers to
#   temp = dn.split('!')
#   # temp[0] is the worksheet name; temp[1] is the cell reference
#   cell = ws[temp[1]]
#   row_index = cell.row
#   column_index = cell.col_idx
#   cell_value = ws.cell(row_index,column_index).value
#
# Get a cell, given a worksheet, and row and column indices
#    cell = ws.cell(163, 24)
#
# Get the fill and fill pattern type of a cell
#    fill = cell.fill
#    patternType = fill.patternType
#
# N.B. The 'magic' fill patternType indicating a filled-in cell in the 
#      schedule exhibit is 'gray125'
#
###############################################################################

import openpyxl
import re

# Fill style of filled-in cells in the schedule exhibit 
MAGIC_FILL_STYLE = 'gray125'

# Return the column index for a defined name assigned to A SINGLE CELL.
# Note: In Excel, the scope of 'defined names' is the entire workBOOK, not a particular workSHEET.
def get_column_index(wb, name):
    ws = wb['workscope_exhibits']
    x = wb.defined_names[name].value
    temp = x.split('!')
    # temp[0] is the worksheet reference, temp[1] is the cell reference
    cell = ws[temp[1]]
    col_ix = cell.col_idx
    return col_ix
# end_def get_column_index()

# Return the row index for a defined name assigned to A SINGLE CELL.
# Note: In Excel, the scope of 'defined names' is the workBOOK, not a particular workSHEET.
def get_row_index(wb, name):
    ws = wb['workscope_exhibits']
    x = wb.defined_names[name].value
    temp = x.split('!')
    # temp[0] is the worksheet reference, temp[1] is the cell reference
    cell = ws[temp[1]]
    row = cell.row
    return row
# end_def get_row_index()

# Return the contents of a cell.
# If OpenPyXl cell accessor raises exception OR value returned by OpenPyXl accessor is None, return the empty string.
def get_cell_contents(ws, row_ix, col_ix):
    try:
        temp = ws.cell(row_ix, col_ix).value
    except:
        temp = ''
    if temp == None:
        retval = ' '
    else:
        retval = temp
    return retval
# end_def get_cell_contents()

# Return the column index of the right-most schedule column
# that is either filled-in as part of a task duration or
# contains an upper-case character indicating a milestone.
# This function is logically nested within initExcelFile,
# but has been coded here at scope-0 for the sake of easy
# development and debugging.
def get_last_used_sched_column(xlsInfo):
    rv = 0
    ws = xlsInfo['ws']
    first_col = xlsInfo['first_schedule_col_ix']   
    last_col = xlsInfo['last_schedule_col_ix']
    first_row = xlsInfo['task_list_top_row_ix']
    last_row = xlsInfo['task_list_bottom_row_ix']
    # Search 'backwards' in time through the schedule, i.e., right-to-left
    for col in range(last_col - 1, first_col -1 , -1):
        bv = ''
        for row in range(first_row+1,last_row):
            cell = ws.cell(row,col)
            fill = cell.fill
            patternType = fill.patternType
            contents = get_cell_contents(ws,row,col)
            if patternType == MAGIC_FILL_STYLE or str(contents).isupper():
                bv += '1'
            else:
                bv += '0'
            # end_if
        # end_for
        if re.search('1',bv) != None:
            # Column has real data
            rv = col
            break
        # end_if
    # end_for
    return rv
# end_def get_last_used_sched_column()

# This should dump the contents of pretty much any dictionary passed to it.
# It is primarily intended, though, to dump 'xlsInfo' during development/debug.
def dump_xlsInfo(xlsInfo):
    l = []
    for item in xlsInfo:
        l.append(item)
    # end_for
    l.sort()
    for item in l:
       print item + ': ' + str(xlsInfo[item])
    # end_for()
# end_def dump_xlsInfo()

# Open the workbook (.xlsx file) inidicated by the "fullpath" parameter.
# Return a dictionary containing the items listed below, which is in
# (almost) alphabetical order. The meaning of most of these entries
# is self-evident from their names, or from consulting the comment
# block above that doucments the 'defined names' which must be present
# in the input .xlsx file. When this is not the case, a description is
# given below.
#
#   errors - string with text of error message(s) for any error(s) 
#            encountered when reading the input .xlsx file.
#            If 'errors' == '', processing found no errors.
#   direct_salary_cell_col_ix
#   direct_salary_cell_row_ix
#   direct_salary_col_ix
#   first_schedule_col_ix
#   funding_list_bottom_row_ix
#   funding_list_top_row_ix
#   funding_source_name_col_ix
#   last_schedule_col_ix
#   last_used_schedule_col_ix - the index of the last data cell in the
#                               schedule table that has been filled-in
#                               by hatching and/or contains a milestone
#   m1_col_ix
#   milestone_label_col_ix
#   milestone_name_col_ix
#   milestones_list_first_row_ix
#   num_sched_col_header_cells - the number of column header cells 
#                                required in the output HTML table
#                                for the schedule in the input .xlsx file
#   num_sched_subdivisions - the number of subdivisions PER COLUMN HEADER
#                            required in the output HTML table; legal
#                            value can only be 3, 4, or 5
#   odc_cell_col_ix
#   odc_cell_row_ix
#   odc_consultants_line_ix
#   odc_dp_equipment_line_ix
#   odc_office_equipment_line_ix
#   odc_other_line_ix
#   odc_printing_line_ix
#   odc_travel_line_ix
#   overhead_cell_col_ix
#   overhead_cell_row_ix
#   overhead_col_ix
#   p1_col_ix
#   p2_col_ix
#   p3_col_ix
#   p4_col_ix
#   p5_col_ix
#   project_name_cell_col_ix
#   project_name_cell_row_ix
#   sched_major_units - the 'major scheduling unit' used in the input .xlsx file;
#                       legal value can only be 'Quarter', 'Month', or 'Week'
#   sched_minor_units - the 'minor scheduling unit' implied by the major 
#                       scheduling unit selected by the user; legal value 
#                       can only be 'Months', 'Weeks', or 'Days'
#   sp1_col_ix
#   sp3_col_ix
#   task_list_bottom_row_ix
#   task_list_top_row_ix
#   task_name_col_ix
#   task_number_col_ix
#   temp_col_ix
#   total_col_ix
#   total_cost_cell_col_ix
#   total_cost_cell_row_ix
#   total_cost_col_ix
#   total_line_row_ix
#   wb - the .xlsx workbook that was opened
#   ws - the 'workscope_exhibits' worksheet
#
def initExcelFile(fullpath):
    # retval dictionary
    retval = {}
    retval['errors'] = ''
    
    # Workbook MUST be opened with the data_only parameter set to True.
    # This ensures that we read the computed value in cells containing a formula, not the formula itself.
    try:
        wb = openpyxl.load_workbook(fullpath, data_only=True)
        retval['wb'] = wb
    except:
        retval['errors'] += 'Failed to open and/or load input .xlsx file.\n'
    # 
    # N.B. The worksheet containing the workscope exhibits is named 'workscope_exhibits'.
    try:
        ws = wb['workscope_exhibits']
        retval['ws'] = ws
    except:
        retval['errors'] += 'Failed to find workscope_exhibits worksheet.\n'
    # Collect row and column indices for cells of interest for Exhibit 2
    #
    try:
        retval['project_name_cell_row_ix'] = get_row_index(wb, 'project_name_cell')
        retval['project_name_cell_col_ix'] = get_column_index(wb, 'project_name_cell')
    except:
        retval['errors'] += 'Failed to find project_name_cell_row_ix and/or project_name_cell_col_ix\n'
    try:
        retval['direct_salary_cell_row_ix'] = get_row_index(wb, 'direct_salary_cell')
        retval['direct_salary_cell_col_ix'] = get_column_index(wb, 'direct_salary_cell')
    except:
        retval['errors'] += 'Failed to find direct_salary_cell_row_ix and/or direct_salary_cell_col_ix\n'
    try:
        retval['odc_cell_row_ix'] = get_row_index(wb, 'odc_cell')
        retval['odc_cell_col_ix'] = get_column_index(wb, 'odc_cell')
    except:
        retval['errors'] += 'Failed to find odc_cell_row_ix and/or odc_cell_col_ix\n'
    try:
        retval['total_cost_cell_row_ix'] = get_row_index(wb, 'total_cost_cell')
        retval['total_cost_cell_col_ix'] = get_column_index(wb, 'total_cost_cell')
    except:
        retval['errors'] += 'Failed to find total_cost_cell_row_ix and/or total_cost_cell_col_ix\n'
    # Overhead rate cell.
    try:
        retval['overhead_cell_row_ix'] = get_row_index(wb, 'overhead_cell')
        retval['overhead_cell_col_ix'] = get_column_index(wb, 'overhead_cell')
    except:
        retval['errors'] += ' Failed to find overhead_cell_row_ix and/or overhead_cell_col_ix\n'
    #       
    # Collect useful row indices for Exhibit 2
    #
    try:
        retval['task_list_top_row_ix'] = get_row_index(wb, 'task_list_top')
    except:
        retval['errors'] += 'Failed to find task_list_top_row_ix\n'
    try:
        retval['task_list_bottom_row_ix'] = get_row_index(wb, 'task_list_bottom')
    except:
        retval['errors'] += 'Failed to find task_list_bottom_row_ix\n'
    try:
        retval['total_line_row_ix'] = get_row_index(wb, 'total_line')   
    except:
        retval['errors'] += 'Failed to find total_line_row_ix\n'
    # Rows containing other direct costs
    try:
        retval['odc_travel_line_ix'] =  get_row_index(wb, 'odc_travel_line')
    except:
        retval['errors'] += 'Failed to find odc_travel_line_ix\n'
    try:
        retval['odc_office_equipment_line_ix'] = get_row_index(wb, 'odc_office_equipment_line')
    except:
        retval['errors'] += 'Failed to find odc_office_equipment_line_ix\n'
    try:
        retval['odc_dp_equipment_line_ix'] = get_row_index(wb, 'odc_dp_equipment_line')
    except:
        retval['errors'] = 'Failed to find odc_dp_equipment_line_ix\n'
    try:
        retval['odc_consultants_line_ix'] = get_row_index(wb, 'odc_consultants_line')
    except:
        retval['errors'] += 'Failed to find odc_consultants_line_ix\n'
    try:
        retval['odc_printing_line_ix'] = get_row_index(wb, 'odc_printing_line')
    except:
        retval['errors'] += 'odc_printing_line_ix\n'
    try:
        retval['odc_other_line_ix'] = get_row_index(wb, 'odc_other_line')   
    except:
        retval['errors'] += 'Failed to find odc_other_line_ix\n'
    # Rows containing info on funding source(s)
    try:
        retval['funding_list_top_row_ix'] = get_row_index(wb, 'funding_list_top')
    except:
        retval['errors'] += 'Failed to find funding_list_top_row_ix\n'
    try:
        retval['funding_list_bottom_row_ix'] = get_row_index(wb, 'funding_list_bottom')
    except:
        retval['errors'] += 'Failed to find funding_list_bottom_row_ix\n'
    #
    # Collect useful column indices for Exhibit 2
    #
    try:
        retval['task_number_col_ix'] = get_column_index(wb, 'task_number_column')
    except:
        retval['errors'] += 'Failed to find task_number_col_ix\n'
    try:
        retval['task_name_col_ix'] = get_column_index(wb, 'task_name_column')
    except:
        retval['errors'] += 'Failed to find task_name_col_ix\n'
    try:
        retval['m1_col_ix'] = get_column_index(wb, 'm1_column')
    except:
        retval['errors'] += 'Failed to find m1_col_ix\n'
    try:
        retval['p5_col_ix'] = get_column_index(wb, 'p5_column')
    except:
        retval['errors'] += 'Failed to find p5_col_ix\n'
    try:    
        retval['p4_col_ix'] = get_column_index(wb, 'p4_column')
    except:
        retval['errors'] += 'Failed to find p4_col_ix\n'
    try:
        retval['p3_col_ix'] = get_column_index(wb, 'p3_column')
    except:
        retval['errors'] += 'Failed to find p3_col_ix\n'
    try:
        retval['p2_col_ix'] = get_column_index(wb, 'p2_column')
    except:
        retval['errors'] += 'Failed to find p2_col_ix\n'
    try:
        retval['p1_col_ix'] = get_column_index(wb, 'p1_column')
    except:
        retval['errors'] += 'Failed to find p1_col_ix\n'
    try:
        retval['sp3_col_ix'] = get_column_index(wb, 'sp3_column')
    except:
        retval['errors'] += 'Failed to find sp3_col_ix\n'
    try:
        retval['sp1_col_ix'] = get_column_index(wb, 'sp1_column')
    except:
        retval['errors'] += 'Failed to find sp1_col_ix\n'
    try:
        retval['temp_col_ix'] = get_column_index(wb, 'temp_column')
    except:
        retval['errors'] += 'Failed to find temp_col_ix\n'
    # The following statement refers to the column for total labor cost before overhead
    try:
        retval['total_col_ix'] = get_column_index(wb, 'total_column')
    except:
        retval['errors'] += 'Failed to find total_col_ix\n'
    try:
        retval['direct_salary_col_ix'] = get_column_index(wb, 'direct_salary_column')
    except:
        retval['errors'] += 'Failed to find direct_salary_col_ix\n'
    try:
        retval['overhead_col_ix'] = get_column_index(wb, 'overhead_column')
    except:
        retval['errors'] += 'Failed to find total_col_ix\n'
    try:
        retval['total_cost_col_ix'] = get_column_index(wb, 'total_cost_column')
    except:
        retval['errors'] += 'Failed to find total_cost_col_ix\n'
    #
    # C'est un petit hacque: The column index for funding source names is the same as that for task names.
    #
    retval['funding_source_name_col_ix'] = retval['task_name_col_ix']
    
    # Collect row and column indices for cells of interest for Exhibit 1
    #
    try:
        retval['first_schedule_col_ix'] = get_column_index(wb, 'first_schedule_column')
    except:
        retval['errors'] += 'Failed to find first_schedule_col_ix\n'
    try:
        retval['last_schedule_col_ix'] = get_column_index(wb, 'last_schedule_column')
    except:
        retval['errors'] += 'Failed to find last_schedule_col_ix\n'
    try:
        retval['milestone_label_col_ix'] = get_column_index(wb, 'milestone_label_column')
    except:
        retval['errors'] += 'Failed to find milestone_label_col_ix\n'
    try:
        retval['milestone_name_col_ix'] = get_column_index(wb, 'milestone_name_column')
    except:
        retval['errors'] += 'Failed to find milestone_name_col_ix\n'
    try:
        retval['milestones_list_first_row_ix'] = get_row_index(wb, 'milestones_list_first_row')
    except:
        retval['errors'] += 'Failed to find milestones_list_first_row_ix\n'
    # N.B. The last row of the milestones list is found programmatically by crawling down
    #      milestone_label_column until the first row containing a blank cell is found.
    
    try:
        row_ix = get_row_index(wb, 'sched_major_units_cell')
        col_ix = get_column_index(wb, 'sched_major_units_cell')
        maj_units = get_cell_contents(ws, row_ix, col_ix)
        retval['sched_major_units'] = maj_units
        if maj_units == 'Quarter':
            min_units = 'Month'
            num_subdivisions = 3
        elif maj_units == 'Month':
            min_units = 'Week'
            num_subdivisions = 4
        else:
            # Assume major unit is 'Weeks'
            min_units = 'Day'
            num_subdivisions = 5
        # end_if
        retval['sched_minor_units'] = min_units
        retval['num_sched_subdivisions'] = num_subdivisions
    except:
        retval['errors'] += 'Failed to find sched_major_units_cell\n'
    #
    # N.B. The slightly incomplete 'retval' is now passed to get_last_used_sched_column
    last_used_schedule_col_ix = get_last_used_sched_column(retval)
    retval['last_used_schedule_col_ix'] = last_used_schedule_col_ix
    num_minor_sched_units = retval['last_used_schedule_col_ix'] - retval['first_schedule_col_ix'] + 1
    
    # Debug
    # print '*** num_minor_sched_units : ' + str(num_minor_sched_units)
    
    num_major_sched_units = (num_minor_sched_units/num_subdivisions)
    num_major_sched_units += 0 if ((num_minor_sched_units % num_subdivisions) == 0) else 1
      
    # Debug      
    # print '*** num_major_sched_units : ' + str(num_major_sched_units)
    
    retval['num_sched_col_header_cells'] = num_major_sched_units
    return retval
# end_def initExcelFile()
